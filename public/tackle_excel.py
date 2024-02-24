#!/usr/bin/env python
# _*_ coding: utf-8 _*_
# @Author: Bruce·li
# @File: tackle_excel.py
# @Time: 2023-02-01 20:21
# @Version：V 0.1
# @Describe: 处理excel文件
import os
import xlrd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, colors, PatternFill, Border, Side

from public.conf import change_dict, settings, UploadPath


class ExcelObj:
    def __init__(self):
        self.name = None
        self.code = None
        self.bar_code = None
        self.cost_price = None
        self.sale_price = None
        self.company = None
        self.buy_number = None
        self.return_number = None
        self.change_number = None
        self.goods_class = None
        self.goods_date = None
        self.produce_date = None
        self.preserve = None
        self.brand = None
        self.split = None
        self.split_code = None
        self.remark = None
        self.describe = None


def splice_path(file_name: str):
    """
    拼接路径
    """
    file_list = os.listdir(UploadPath)
    for file in file_list:
        name_list = file.split(".")
        if file == file_name or file_name == name_list[0]:
            file_path = os.path.join(UploadPath, file)
            return file_path


class XlsxExcel:
    """
    处理xlsx文件
    """

    def __init__(self, file_name: str = "", sheet_name: str = "data"):
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.wb = Workbook()
        self.ws = self.wb.create_sheet(title=self.sheet_name, index=0)
        self.ws = self.wb.active

    def read_excel(self) -> list:
        """
        读取excel，返回excel对象列表
        """
        file_path = splice_path(self.file_name)
        file_obj = load_workbook(file_path)
        try:
            sheet = file_obj.get_sheet_by_name(self.sheet_name)
        except KeyError:
            return []
        obj_list = list()
        for row in range(2, sheet.max_row + 1):
            obj = ExcelObj()
            for col in range(1, sheet.max_column + 1):
                key = sheet.cell(1, col).value
                value = sheet.cell(row, col).value
                if key in change_dict.keys():
                    setattr(obj, change_dict[key], value)
            obj_list.append(obj)
        return obj_list

    def write(self, row_n, col_n, value, flag=0):
        """写入数据，如(2,3，"hello"),第二行第三列写入数据"hello\""""
        ft = Font(size=10, name='微软雅黑')
        alignment_center = Alignment(horizontal='center', vertical='center')
        if row_n == 1:
            pa = PatternFill("solid", fgColor="CCCC99")
            al = Alignment(horizontal='left', vertical='center', wrap_text=True)
            thin = Side(border_style="thin", color="989898")  # 边框样式，颜色
            border = Border(left=thin, right=thin, top=thin, bottom=thin)  # 边框的位置
            self.ws.cell(row_n, col_n).font = ft
            self.ws.cell(row_n, col_n).fill = pa
            self.ws.cell(row_n, col_n).alignment = al
            self.ws.cell(row_n, col_n).border = border
        else:
            self.ws.cell(row_n, col_n).font = ft
        if flag == 1:  # 入库单表头
            self.ws.row_dimensions[row_n].height = 40
            self.ws.merge_cells(range_string="A1:J1")
        elif flag == 2:  # 时间
            self.ws.row_dimensions[row_n].height = 40
            self.ws.merge_cells(range_string="B2:C2")
        elif flag == 3:  # 供应商
            self.ws.row_dimensions[row_n].height = 40
            self.ws.merge_cells(range_string="E2:G2")
        elif flag == 4:  # 入库单号
            self.ws.row_dimensions[row_n].height = 40
            self.ws.merge_cells(range_string="I2:J2")
        elif flag == 5:  # 出库单表头
            self.ws.row_dimensions[row_n].height = 40
            self.ws.merge_cells(range_string="A1:H1")
        elif flag == 6:  # 出库单号
            self.ws.row_dimensions[row_n].height = 40
            self.ws.merge_cells(range_string="G2:H2")
        elif flag == 7:
            ft = Font(size=10, name='微软雅黑', color="FF0000")
            self.ws.cell(row_n, col_n).font = ft
        else:
            self.ws.row_dimensions[row_n].height = 20
        self.ws.column_dimensions["B"].width = 30
        self.ws.column_dimensions["C"].width = 15
        self.ws.column_dimensions["D"].width = 15
        self.ws.cell(row_n, col_n, value=value).alignment = alignment_center
        self.wb.save(self.file_name)


class XlsExcel:
    """
    处理xls文件
    """

    def __init__(self, file_name: str = "", sheet_name: str = "Sheet1"):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def read_excel(self) -> list:
        """
        读取excel，返回excel对象列表
        """
        file_path = splice_path(self.file_name)
        data = xlrd.open_workbook(file_path)
        try:
            table = data.sheet_by_name(self.sheet_name)
        except KeyError:
            return []
        keys = table.row_values(0)  # 获取第一行作为key值
        rowNum = table.nrows  # 获取总行数
        colNum = table.ncols  # 获取总列数
        obj_list = list()
        j = 1
        for i in list(range(rowNum - 1)):  # 去掉行首 self.rowNum - 1
            obj = ExcelObj()
            values = table.row_values(j)  # 从第二行取对应values值
            for x in list(range(colNum)):
                if keys[x] in change_dict.keys():
                    setattr(obj, change_dict[keys[x]], values[x])
            obj_list.append(obj)
            j += 1
        return obj_list  # 返回list包含的dict数据
